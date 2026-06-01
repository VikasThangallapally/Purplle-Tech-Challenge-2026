from openpyxl import load_workbook
wb = load_workbook('Brigade Road - Store layoutc5f5d56.xlsx', read_only=True)
print('sheets=', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\nSheet:', name)
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        print(row)
        if i>10:
            break
